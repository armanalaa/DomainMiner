"""
build_tune_params_results.py
----------------------------
Place this script inside the dataset folder (e.g. Northwind/).
Scans ccm_output/ for all tA*_tT*_r* run folders and writes:
  - ccm_output/tune_params_results.xlsx  (formatted Excel)
  - ccm_output/tune_params_summary.txt   (plain text summary)

Run from inside the dataset folder:
    python build_tune_params_results.py
"""

import os, re, csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ── Folder name parser ────────────────────────────────────────────────────────

def parse_folder(name):
    m = re.match(r"tA([\d.]+)_tT([\d.]+)_r([\d.]+)$", name)
    if not m:
        return None
    return float(m.group(1)), float(m.group(2)), float(m.group(3))


# ── Report parsers ────────────────────────────────────────────────────────────

def read(path):
    return open(path, encoding="utf-8", errors="ignore").read() if os.path.exists(path) else ""

def get(pattern, text, cast=str):
    m = re.search(pattern, text)
    return cast(m.group(1)) if m else None

def parse_step3(path):
    t = read(path)
    return {
        "GA_edges":   get(r"Edges in G_A\s*=\s*(\d+)", t, int),
        "pruned_pct": get(r"Pruned pairs\s*=\s*\d+\s*\(([\d.]+)%\)", t, float),
    }

def parse_step4(path):
    t = read(path)
    return {
        "table_pairs": get(r"Table pairs computed\s*:\s*(\d+)", t, int),
        "GT_edges":    get(r"Edges in G_T\s*:\s*(\d+)", t, int),
    }

def parse_step5(path):
    t = read(path)
    tables    = get(r"Tables in G_T\s*:\s*(\d+)", t, int)
    domains   = get(r"Domains discovered\s*:\s*(\d+)", t, int)
    q_raw     = get(r"Modularity Q\s*:\s*([\d.]+)", t, float)
    q         = round(q_raw, 4) if q_raw else None
    valid     = "YES" if q and q >= 0.3 else "NO"

    blocks = re.findall(r"D\d+\s+(.+?)\n.*?Tables\s*:\s*(.+?)(?:\n|$)", t, re.DOTALL)
    names, sizes = [], []
    for name, tstr in blocks:
        tbls = [x.strip() for x in tstr.split(",") if x.strip()]
        names.append(name.strip())
        sizes.append(len(tbls))

    return {
        "tables_in_GT": tables,
        "domains":      domains,
        "Q":            q,
        "Q_valid":      valid,
        "singletons":   sum(1 for s in sizes if s == 1) if sizes else None,
        "largest_dom":  max(sizes) if sizes else None,
        "domain_names": " | ".join(names) if names else None,
        "domain_sizes": " | ".join(str(s) for s in sizes) if sizes else None,
    }


# ── Excel writer ──────────────────────────────────────────────────────────────

NAVY        = "1F3864"
NAVY_LIGHT  = "2F5496"
HEADER_BG   = "1F3864"
WHITE       = "FFFFFF"
YELLOW      = "FFFF00"
GREEN_BG    = "E2EFDA"
RED_BG      = "FCE4D6"

COLS = [
    ("theta_A",    "theta_a",    8),
    ("theta_T",    "theta_t",    8),
    ("Resolution", "resolution", 10),
    ("Run Tag",    "folder",     30),
    ("Tables",     "tables_in_GT", 8),
    ("Edges G_T",  "GT_edges",   10),
    ("Domains",    "domains",    9),
    ("Q",          "Q",          9),
    ("Status",     "Q_valid",    9),
    ("Time (s)",   None,         10),   # placeholder — not in reports
]

def write_xlsx(rows, path, dataset_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ncols = len(COLS)
    ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Row 1: title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1,
        value=f"CCM Parameter Tuning — {dataset_name}   |   Generated: {ts}")
    title_cell.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
    title_cell.fill      = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Row 2: headers
    for col_idx, (header, _, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 16

    # Data rows
    for row_idx, r in enumerate(rows, start=3):
        q     = r.get("Q")
        valid = r.get("Q_valid", "NO")
        bg    = GREEN_BG if valid == "YES" else RED_BG

        for col_idx, (_, key, _) in enumerate(COLS, start=1):
            val  = r.get(key) if key else None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")

            # Bold Q value
            if key == "Q" and val is not None:
                cell.font = Font(name="Arial", size=10, bold=True)
            # Status cell color text
            if key == "Q_valid":
                color = "375623" if valid == "YES" else "9C0006"
                cell.font = Font(name="Arial", size=10, bold=True, color=color)

    # Freeze header rows
    ws.freeze_panes = "A3"

    wb.save(path)


# ── Summary TXT ───────────────────────────────────────────────────────────────

def write_txt(rows, path):
    valid = [r for r in rows if r["Q_valid"] == "YES"]
    with open(path, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("tune_params — Results Summary\n")
        f.write("=" * 70 + "\n\n")
        f.write(f"Total runs    : {len(rows)}\n")
        f.write(f"Valid (Q>=0.3) : {len(valid)}\n\n")
        f.write(f"{'Folder':<28} {'tA':>5} {'tT':>5} {'res':>5} "
                f"{'GT_e':>6} {'dom':>5} {'Q':>7} {'valid':>6}\n")
        f.write("-" * 75 + "\n")
        for r in rows:
            q_str = f"{r['Q']:.4f}" if r["Q"] is not None else "  N/A"
            f.write(f"{r['folder']:<28} {r['theta_a']:>5} {r['theta_t']:>5} "
                    f"{r['resolution']:>5} {str(r['GT_edges']):>6} "
                    f"{str(r['domains']):>5} {q_str:>7} {r['Q_valid']:>6}\n")
        if valid:
            best = valid[0]
            f.write("\n" + "=" * 70 + "\n")
            f.write("Best run\n")
            f.write("=" * 70 + "\n")
            for k in ["folder","theta_a","theta_t","resolution",
                      "Q","domains","singletons","domain_names","domain_sizes"]:
                f.write(f"  {k:<15}: {best.get(k)}\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    dataset_dir  = os.path.dirname(os.path.abspath(__file__))
    dataset_name = os.path.basename(dataset_dir)
    ccm_dir      = os.path.join(dataset_dir, "ccm_output")

    if not os.path.isdir(ccm_dir):
        print(f"ERROR: ccm_output/ not found in {dataset_dir}")
        return

    run_folders = sorted([
        d for d in os.listdir(ccm_dir)
        if os.path.isdir(os.path.join(ccm_dir, d)) and parse_folder(d)
    ])

    if not run_folders:
        print("No run folders found in ccm_output/")
        return

    print(f"Found {len(run_folders)} runs in {dataset_name}/ccm_output/\n")

    rows = []
    for folder in run_folders:
        ta, tt, res = parse_folder(folder)
        run_dir = os.path.join(ccm_dir, folder)
        s3 = parse_step3(os.path.join(run_dir, "step3_sim_attr_report.txt"))
        s4 = parse_step4(os.path.join(run_dir, "step4_report.txt"))
        s5 = parse_step5(os.path.join(run_dir, "step5_report.txt"))

        row = {"folder": folder, "theta_a": ta, "theta_t": tt, "resolution": res,
               **s3, **s4, **s5}
        rows.append(row)

        q    = s5["Q"]
        mark = s5["Q_valid"]
        q_str = f"{q:.4f}" if q else "N/A"
        print(f"  {folder:<28}  Q={q_str}  {mark}  domains={s5['domains']}")

    rows.sort(key=lambda r: r["Q"] if r["Q"] else -1, reverse=True)

    xlsx_path = os.path.join(ccm_dir, "tune_params_results.xlsx")
    txt_path  = os.path.join(ccm_dir, "tune_params_summary.txt")

    write_xlsx(rows, xlsx_path, dataset_name)
    write_txt(rows, txt_path)

    valid = [r for r in rows if r["Q_valid"] == "YES"]
    print(f"\nExcel   -> ccm_output/tune_params_results.xlsx")
    print(f"Summary -> ccm_output/tune_params_summary.txt")
    print(f"\nTotal: {len(rows)}  |  Valid: {len(valid)}")
    if valid:
        best = valid[0]
        print(f"Best : {best['folder']}  Q={best['Q']}  domains={best['domains']}")
        print(f"Names: {best['domain_names']}")


if __name__ == "__main__":
    main()