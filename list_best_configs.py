"""
list_best_configs.py
====================
Reads the first row (best Q) from each dataset's tune_params_results.xlsx,
reads domain names from the corresponding step5_domains.json, and writes
a single summary Excel file to the project root.

USAGE
-----
  python list_best_configs.py
  python list_best_configs.py --datasets Sakila Northwind Chinook
  python list_best_configs.py --output results/best_configs.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# Configuration
# =============================================================================

KNOWN_DATASETS = [
    "Sakila", "Northwind", "Chinook", "DellStore2", "adventure_works",
    "WideWorldImporters", "Employees", "TPCDS", "Airportdb",
    "FDA_AdverseEvents", "StackOverflowDataDump", "eicu", "mimiciv",
    "Synthea", "tcph",
]

RUN_PATTERN = re.compile(r"tA([\d.]+)_tT([\d.]+)_r([\d.]+)")

# =============================================================================
# Read dataset stats from schema.json
# =============================================================================

def read_dataset_stats(dataset_dir: Path) -> dict:
    """Read total tables and columns from schema.json; count rows from CSV files."""
    schema_path = dataset_dir / "schema.json"
    total_tables  = None
    total_columns = 0
    total_rows    = 0

    # Tables and columns from schema.json
    if schema_path.exists():
        try:
            data   = json.loads(schema_path.read_text(encoding="utf-8", errors="ignore"))
            tables = data.get("tables", {})
            if isinstance(tables, list):
                tables = {t.get("name", str(i)): t for i, t in enumerate(tables)}
            total_tables = len(tables)
            for tdata in tables.values():
                if isinstance(tdata, dict):
                    total_columns += len(tdata.get("columns", []))
        except Exception as e:
            print(f"  [WARN] schema.json unreadable for {dataset_dir.name}: {e}")

    # Row counts from CSV files
    csv_dir = dataset_dir / "csv"
    if csv_dir.exists():
        for csv_file in csv_dir.glob("*.csv"):
            try:
                with open(csv_file, encoding="utf-8", errors="ignore") as f:
                    # Count lines minus header
                    total_rows += max(0, sum(1 for _ in f) - 1)
            except Exception:
                pass

    return {
        "total_tables":  total_tables,
        "total_columns": total_columns if total_columns > 0 else None,
        "total_rows":    total_rows    if total_rows    > 0 else None,
    }

# =============================================================================

def read_best_from_xlsx(dataset_dir: Path) -> dict | None:
    """Read the first data row (highest Q) from tune_params_results.xlsx."""
    xlsx_path = dataset_dir / "ccm_output" / "tune_params_results.xlsx"
    if not xlsx_path.exists():
        return None
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        # Row 1 = title, Row 2 = headers, Row 3 = best run (sorted by Q desc)
        row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        if row[0] is None:
            return None
        return {
            "theta_a":    row[0],
            "theta_t":    row[1],
            "resolution": row[2],
            "run_tag":    row[3],
            "n_tables":   row[4],
            "n_edges":    row[5],
            "n_domains":  row[6],
            "Q":          row[7],
            "status":     row[8],
            "elapsed_s":  row[9],
        }
    except Exception as e:
        print(f"  [WARN] Could not read {xlsx_path}: {e}")
        return None

# =============================================================================
# Read domain names from step5_domains.json of the best run folder
# =============================================================================

def read_domain_names(dataset_dir: Path, run_tag: str) -> list[str]:
    """Read domain names from the best run's step5_domains.json."""
    domains_path = dataset_dir / "ccm_output" / run_tag / "step5_domains.json"
    if not domains_path.exists():
        return []
    try:
        data = json.loads(domains_path.read_text(encoding="utf-8", errors="ignore"))
        if isinstance(data, list):
            return [
                str(d.get("domain_name") or d.get("name") or "").strip()
                for d in data if isinstance(d, dict)
            ]
    except Exception:
        pass
    return []

# =============================================================================
# Build summary
# =============================================================================

def build_summary(root: Path, datasets: list[str]) -> list[dict]:
    records = []
    for name in datasets:
        dataset_dir = root / name
        if not dataset_dir.is_dir():
            continue

        best = read_best_from_xlsx(dataset_dir)
        if best is None:
            print(f"  [SKIP] {name:30s} — tune_params_results.xlsx not found or empty")
            continue

        run_tag = str(best.get("run_tag", ""))
        domain_names = read_domain_names(dataset_dir, run_tag) if run_tag else []
        stats = read_dataset_stats(dataset_dir)

        record = {"Dataset": name}
        record.update(best)
        record["domain_names"] = domain_names
        record.update(stats)

        print(f"  [OK]   {name:30s}  Q={best['Q']:.4f}  "
              f"domains={best['n_domains']}  tag={run_tag}"
              f"  tables={stats['total_tables']}  cols={stats['total_columns']}"
              f"  rows={stats['total_rows']}")

        records.append(record)
    return records

# =============================================================================
# Write Excel
# =============================================================================

def write_excel(records: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Best Configurations"

    dark_blue  = "1F3864"
    mid_blue   = "2F5496"
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    good_fill  = PatternFill("solid", fgColor="EBF3E8")
    alt_fill   = PatternFill("solid", fgColor="EEF3FA")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin       = Side(style="thin", color="BFBFBF")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    left       = Alignment(horizontal="left",   vertical="center")

    # Title
    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value     = (f"DomainDiscover — Best Configuration per Dataset   |   "
                   f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    t.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=dark_blue)
    t.alignment = center
    ws.row_dimensions[1].height = 22

    # Headers
    headers = [
        ("Dataset",            18), ("theta_A",       9), ("theta_T",     9),
        ("Resolution",         11), ("Run Tag",       24),
        ("Total Tables",       12), ("Total Columns", 13), ("Total Rows",  14),
        ("Tables G_T",          9), ("Edges G_T",     10), ("Domains",     9),
        ("Q",                   9), ("Status",        10), ("Time (s)",    10),
        ("Discovered Domains", 70),
    ]
    for col, (label, width) in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=mid_blue)
        c.alignment = center
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 16

    # Data rows sorted by Q descending
    records_sorted = sorted(records, key=lambda r: r.get("Q") or 0, reverse=True)
    best_q = records_sorted[0]["Q"] if records_sorted else None

    for i, rec in enumerate(records_sorted):
        row    = i + 3
        q      = rec.get("Q")
        is_best = q == best_q
        fill   = (green_fill if is_best else
                  good_fill  if (q and q >= 0.3) else
                  alt_fill   if i % 2 == 0 else white_fill)

        def dc(col, value, fmt=None, bold=False, wrap=False):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(name="Arial", size=10 if col < 12 else 9, bold=bold)
            c.fill      = fill
            c.border    = border
            c.alignment = (Alignment(horizontal="left", vertical="center",
                                     wrap_text=wrap)
                           if col in (1, 12) else center)
            if fmt:
                c.number_format = fmt

        dc(1,  rec["Dataset"],    bold=is_best)
        dc(2,  rec["theta_a"])
        dc(3,  rec["theta_t"])
        dc(4,  rec["resolution"])
        dc(5,  rec["run_tag"])
        dc(6,  rec.get("total_tables"))
        dc(7,  rec.get("total_columns"))
        dc(8,  rec.get("total_rows"), fmt="#,##0")
        dc(9,  rec["n_tables"])
        dc(10, rec["n_edges"])
        dc(11, rec["n_domains"])
        dc(12, round(q, 4) if q is not None else None,
           fmt="0.0000", bold=is_best)
        dc(13, rec["status"])
        dc(14, rec["elapsed_s"])

        names      = rec.get("domain_names", [])
        domain_str = "  |  ".join(f"D{i}: {n}" for i, n in enumerate(names))
        dc(15, domain_str, wrap=True)

        ws.row_dimensions[row].height = max(15, 14 * ((len(domain_str) // 70) + 1))

    # Legend
    lr = len(records_sorted) + 4
    ws.merge_cells(f"A{lr}:O{lr}")
    leg = ws[f"A{lr}"]
    leg.value = ("Colour scale:  Green = highest Q overall   |   "
                 "Light green = Q ≥ 0.3 (valid)   |   Blue = Q < 0.3")
    leg.font      = Font(name="Arial", italic=True, size=9, color="595959")
    leg.alignment = left

    ws.freeze_panes = "A3"
    wb.save(output_path)
    print(f"\n  Saved: {output_path}")

# =============================================================================
# Main
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="List best configuration per dataset with domain names."
    )
    parser.add_argument("--output",   type=Path, default=Path("best_configs_summary.xlsx"))
    parser.add_argument("--datasets", nargs="+", default=None)
    parser.add_argument("--root",     type=Path, default=Path("."))
    args = parser.parse_args()

    root     = args.root.resolve()
    datasets = args.datasets or KNOWN_DATASETS

    print(f"\nDomainDiscover — Best Configuration per Dataset")
    print(f"Root   : {root}")
    print(f"Output : {args.output}\n")

    records = build_summary(root, datasets)
    if not records:
        print("\n[ERROR] No results found.")
        return

    print(f"\n  {len(records)} dataset(s) found.\n")
    write_excel(records, args.output)


if __name__ == "__main__":
    main()