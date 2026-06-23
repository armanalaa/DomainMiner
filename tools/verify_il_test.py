from openpyxl import load_workbook


path = r"H:\Other computers\My PC\Sapienza\CAISe2025_Revised\DL-To-DM\python\DomainMiner\outputs\Il Test.xlsx"
wb = load_workbook(path)
ws = wb["Modello_Test"]

print("sheets", wb.sheetnames)
print("tables", list(ws.tables.keys()))
print("dims", ws.max_row, ws.max_column)
print("headers", [ws.cell(1, c).value for c in range(1, 11)])

rows = []
for r in range(2, 7):
    row = [ws.cell(r, c).value for c in range(1, 11)]
    rows.append(row)
    print(
        r,
        row[0],
        row[6],
        row[8],
        row[9],
        "newline",
        "\n" in row[1],
        "literal",
        "\\n" in row[1],
        "endscolon",
        row[1].endswith(":"),
    )

print("corrects", [row[6] for row in rows])
print("difficulties", [row[9] for row in rows])
print("paragraphs", sorted(set(row[8] for row in rows)))
print("fill B2", ws["B2"].fill.fgColor.rgb, "wrap", ws["B2"].alignment.wrap_text)
